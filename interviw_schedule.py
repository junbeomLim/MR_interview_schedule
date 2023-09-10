from openpyxl import load_workbook
import copy
import random

#설문지 엑셀 시트 이름
load_wb = load_workbook('MR Application form.xlsx')

#사용할 시트 이름 
ws = load_wb['설문지 응답 시트1']

#=================================================================#

#면접자가 포함된 row의 범위 지정
row_min = 2
row_max = ws.max_row

#면접 불가능한 시간에 관한 질문이 입력되는 col의 범위 지정
start_time_col = 'N'
end_time_col = 'U'

#면접자 이름이 입력되는 col
name_col = 'C'

#면접을 제일 먼저 시작하는 시간
start_time = 1900 #19:00
#면접 시간 간격
interview_interval = 30

time_line = {}
monday_dict = {}
tuesday_dict = {}
#wednesday_dict = {}

#면접 시간 생성 함수
#==================================================================#

def Create_interview_start_time(first_start_time, time_interval):
    time = first_start_time-time_interval
    for j in range(ord(start_time_col), ord(end_time_col)+1):
        time = time + time_interval
        if time%100 >= 60:
            time = time -60 + 100
        
        time_line[j] = time
    return time_line

#===================================================================#

time_line = Create_interview_start_time(start_time, interview_interval)

for time in time_line.values():
    monday_dict[time] = []
    tuesday_dict[time] = []
    #wednesday_dict[time] = []

time_values = list(time_line.values())
monday = copy.deepcopy(time_values)
tuesday = copy.deepcopy(time_values)
wednesday = copy.deepcopy(time_values)

for row in range(row_min, row_max+1):
    name = ws[name_col+str(row)].value
    
    for col in range(ord(start_time_col), ord(end_time_col)+1):
        day = ws[chr(col)+str(row)].value
        if day != None:
            if '월' in day:
                if time_line[col] in monday:
                    monday.remove(time_line[col])
            
            if '화' in day:
                if time_line[col] in tuesday:
                    tuesday.remove(time_line[col])
            """
            if '수' in day:
                if time_line[col] in wednesday:
                    wednesday.remove(time_line[col])
                cases[name]['수'].remove(time_line[col])
            """

    for time in monday:
        monday_dict[time].append(name)
    for time in tuesday:
        tuesday_dict[time].append(name)

    monday = copy.deepcopy(time_values)
    tuesday = copy.deepcopy(time_values)
    """
    wednesday = copy.deepcopy(time_values)
    """

#monday
chosed_people = []
choose_people = []

for time in time_values:
    for chosed_person in chosed_people:
        if chosed_person in monday_dict[time]:
            monday_dict[time].remove(chosed_person)

    if len(monday_dict[time]) > 3:
        choose_people = random.sample(monday_dict[time], 3)
        for chosed_person in choose_people:
            chosed_people.append(chosed_person)
    else:
        choose_people = monday_dict[time]
        for chosed_person in choose_people:
            chosed_people.append(chosed_person)

    monday_dict[time] = copy.deepcopy(choose_people)

#tuesday
for time in time_values:
    for chosed_person in chosed_people:
        if chosed_person in tuesday_dict[time]:
            tuesday_dict[time].remove(chosed_person)

    if len(tuesday_dict[time]) > 3:
        choose_people = random.sample(tuesday_dict[time], 3)
        for chosed_person in choose_people:
            chosed_people.append(chosed_person)
    else:
        choose_people = tuesday_dict[time]
        for chosed_person in choose_people:
            chosed_people.append(chosed_person)

    tuesday_dict[time] = copy.deepcopy(choose_people)

"""
#wednesday
for time in time_values:
    for chosed_person in chosed_people:
        if chosed_person in wednesday_dict[time]:
            wednesday_dict[time].remove(chosed_person)

    if len(wednesday_dict[time]) > 3:
        choose_people = random.sample(tuesday_dict[time], 3)
        for chosed_person in choose_people:
            chosed_people.append(chosed_person)
    else:
        choose_people = wednesday_dict[time]
        for chosed_person in choose_people:
            chosed_people.append(chosed_person)

    wednesday_dict[time] = copy.deepcopy(choose_people)
"""

print("Mon")
for time in monday_dict.keys():
    print(time, *monday_dict[time],'\n')

print("Tues")
for time in tuesday_dict.keys():
    print(time, *tuesday_dict[time],'\n')

"""
print("Wed")
for time in wednesday_dict.keys():
    print(time, *wednesday_dict[time],'\n')
"""

remain = []
for row in range(row_min, row_max+1):
    name = ws['C'+str(row)].value
    if name not in chosed_people:
        remain.append(name)

print("remain: " ,remain)